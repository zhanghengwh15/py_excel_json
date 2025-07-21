#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 兼容性处理：如果python3不可用，尝试使用python
"""
JSON转Excel工具
功能：解析JSON数据，提取菜单信息并转换为Excel格式
作者：AI Assistant
日期：2024
"""

import json
import pandas as pd
from datetime import datetime
import os
import sys
import requests
from typing import List, Dict, Any


class JsonToExcelConverter:
    """JSON转Excel转换器"""
    
    def __init__(self):
        self.menu_data = []
        self.user_menu_data = []  # 新增：存储用户菜单数据
    
    def get_user_list(self, base_url: str = "https://cloudsy.shede.com.cn", 
                     headers: Dict[str, str] = None, 
                     request_data: Dict[str, Any] = None) -> List[Dict[str, Any]]:
        """
        获取用户列表（支持分页）
        
        Args:
            base_url: 基础URL
            headers: 请求头
            request_data: 请求数据
            
        Returns:
            List[Dict]: 用户列表，包含 nickName 和 userId
        """
        import time
        
        url = f"{base_url}/api/poit-cloud-platform/user/ent/user/page"
        
        # 默认请求头
        default_headers = {
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'zh-cn',
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'Content-Type': 'application/json; charset=UTF-8',
            'Origin': base_url,
            'Pragma': 'no-cache',
            'Referer': f'{base_url}/cloud/cloud_config/userManagement',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-origin',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36',
            'X-Requested-With': 'XMLHttpRequest',
            'loglevel': 'debug',
            'menusUri': '/userManagement',
            'orgId': '1000879',
            'poit-cloud-org': '1d7d84a6f6b14d6d97f9c7a94813bb22',
            'poit-cloud-src-client': 'cloud',
            'sec-ch-ua': '"Not)A;Brand";v="8", "Chromium";v="138", "Google Chrome";v="138"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"macOS"',
            'token': 'a7ab82cde0b64c45805d92fa65612c5f',
            'userId': '501073',
            'x-poit-tif-submit': '1675135120_1753070008803'
        }
        
        # 合并自定义请求头
        if headers:
            default_headers.update(headers)
        
        # 默认请求数据
        default_data = {
            "eid": "1d7d84a6f6b14d6d97f9c7a94813bb22",
            "operateUserId": "501073",
            "orgId": "1000879",
            "uid": "77c6e33e1b7d4371aecc6477322ff759",
            "appVersion": "1.0",
            "keyword": "",
            "combineStatus": None,
            "pageSize": 100,
            "pageNum": 1,
            "positionIdList": [],
            "roleIdList": [],
            "factoryIdList": [],
            "departmentIdList": []
        }
        
        # 合并自定义请求数据
        if request_data:
            default_data.update(request_data)
        
        all_users = []
        page_num = 1
        
        try:
            while True:
                # 更新页码
                default_data['pageNum'] = page_num
                
                print(f"📄 正在获取第 {page_num} 页用户数据...")
                
                response = requests.post(url, headers=default_headers, json=default_data)
                response.raise_for_status()
                
                result = response.json()
                
                # 检查响应数据
                if 'data' not in result:
                    print("❌ 响应数据格式不正确，缺少data字段")
                    break
                
                # 如果data为空，说明已经获取完所有数据
                if not result['data'] or len(result['data']) == 0:
                    print(f"✅ 已获取完所有用户数据，共 {len(all_users)} 个用户")
                    break
                
                # 提取用户数据
                if isinstance(result['data'], list):
                    page_users = []
                    for user in result['data']:
                        page_users.append({
                            'nickName': user.get('nickName', ''),
                            'userId': user.get('userId', ''),
                            'openUserId': user.get('openUserId', '')  # 新增：采集openUserId
                        })
                    
                    all_users.extend(page_users)
                    print(f"   ✅ 第 {page_num} 页获取到 {len(page_users)} 个用户")
                    
                    # 如果当前页的用户数量小于pageSize，说明已经是最后一页
                    if len(page_users) < default_data['pageSize']:
                        print(f"✅ 已获取完所有用户数据，共 {len(all_users)} 个用户")
                        break
                else:
                    print("❌ 响应data字段不是数组格式")
                    break
                
                # 页码递增
                page_num += 1
                
                # 请求间隔200ms
                time.sleep(0.2)
                
        except requests.exceptions.RequestException as e:
            print(f"❌ 获取用户列表失败: {str(e)}")
            return all_users  # 返回已获取的数据
        
        return all_users
    
    def get_user_menu_permissions(self, user_id: str, open_user_id: str = None, base_url: str = "https://cloudsy.shede.com.cn",
                                headers: Dict[str, str] = None, 
                                request_data: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        获取指定用户的菜单权限
        
        Args:
            user_id: 用户ID
            base_url: 基础URL
            headers: 请求头
            request_data: 请求数据
            
        Returns:
            Dict: 用户菜单权限数据
        """
        import time
        
        url = f"{base_url}/api/poit-cloud-platform/resources/userId/noWithGeneralize/tree"
        
        # 默认请求头
        default_headers = {
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'zh-cn',
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'Content-Type': 'application/json; charset=UTF-8',
            'Origin': base_url,
            'Pragma': 'no-cache',
            'Referer': f'{base_url}/cloud/userManagement',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-origin',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36',
            'X-Requested-With': 'XMLHttpRequest',
            'loglevel': 'debug',
            'menusUri': '/userManagement',
            'orgId': '1000879',
            'poit-cloud-org': '1d7d84a6f6b14d6d97f9c7a94813bb22',
            'poit-cloud-src-client': 'cloud',
            'sec-ch-ua': '"Not)A;Brand";v="8", "Chromium";v="138", "Google Chrome";v="138"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"macOS"',
            'token': 'a7ab82cde0b64c45805d92fa65612c5f',
            'userId': '501073',
            'x-poit-tif-submit': '5385307557_1753071630108'
        }
        
        # 合并自定义请求头
        if headers:
            default_headers.update(headers)
        
        # 默认请求数据，将 operateUserId 设置为传入的 user_id
        default_data = {
            "eid": "1d7d84a6f6b14d6d97f9c7a94813bb22",
            "operateUserId": user_id,  # 使用传入的用户ID
            "orgId": "1000879",
            "uid": open_user_id if open_user_id else "77c6e33e1b7d4371aecc6477322ff759",  # 使用传入的openUserId
            "appVersion": "1.0"
        }
        
        # 合并自定义请求数据
        if request_data:
            default_data.update(request_data)
            # 确保 operateUserId 使用传入的 user_id
            default_data['operateUserId'] = user_id
        
        try:
            # 请求间隔300ms
            time.sleep(0.3)
            
            response = requests.post(url, headers=default_headers, json=default_data)
            response.raise_for_status()
            
            result = response.json()
            
            # 返回 data 字段中的数组数据
            if 'data' in result:
                return result['data']
            else:
                print(f"❌ 用户 {user_id} 的菜单权限数据格式不正确")
                return []
                
        except requests.exceptions.RequestException as e:
            print(f"❌ 获取用户 {user_id} 菜单权限失败: {str(e)}")
            return []
    
    def parse_user_menu_data(self, user_name: str, menu_data: List[Dict[str, Any]], 
                           parent_path: str = "", level: int = 1) -> None:
        """
        解析用户菜单数据并添加到用户菜单数据列表
        
        Args:
            user_name: 用户名称
            menu_data: 菜单数据列表
            parent_path: 父级菜单路径
            level: 当前层级
        """
        for item in menu_data:
            # 获取显示名称和类型
            display_name = item.get('resourcesDisplayName', '')
            resources_type = item.get('resourcesType', 1)
            
            # 构建菜单全路径
            if parent_path:
                menu_path = f"{parent_path} - {display_name}"
            else:
                menu_path = display_name
            
            # 确定类型显示
            type_display = "菜单" if resources_type == 1 else "功能"
            
            # 添加到用户菜单数据列表
            self.user_menu_data.append({
                '用户名称': user_name,
                '菜单名称': menu_path,
                '菜单层级': level,
                '类型': type_display
            })
            
            # 递归处理子菜单
            child_resources = item.get('childResources', [])
            if child_resources:
                self.parse_user_menu_data(user_name, child_resources, menu_path, level + 1)
    
    def batch_process_user_permissions(self, headers: Dict[str, str] = None,
                                     user_request_data: Dict[str, Any] = None,
                                     menu_request_data: Dict[str, Any] = None) -> bool:
        """
        批量处理用户菜单权限
        
        Args:
            headers: 自定义请求头
            user_request_data: 用户列表请求的自定义数据
            menu_request_data: 菜单权限请求的自定义数据
            
        Returns:
            bool: 处理是否成功
        """
        print("🚀 开始批量处理用户菜单权限...")
        
        # 1. 获取用户列表
        print("📋 正在获取用户列表...")
        users = self.get_user_list(headers=headers, request_data=user_request_data)
        
        if not users:
            print("❌ 未获取到用户数据")
            return False
        
        print(f"✅ 获取到 {len(users)} 个用户")
        
        # 2. 清空之前的数据
        self.user_menu_data = []
        
        # 3. 批量获取每个用户的菜单权限
        for i, user in enumerate(users, 1):
            user_id = user['userId']
            nick_name = user['nickName']
            open_user_id = user.get('openUserId', '')  # 获取openUserId
            
            print(f"📊 正在处理用户 {i}/{len(users)}: {nick_name} (ID: {user_id}, OpenID: {open_user_id})")
            
            # 获取用户菜单权限（内部已有300ms间隔）
            menu_data = self.get_user_menu_permissions(user_id, open_user_id, headers=headers, request_data=menu_request_data)
            
            if menu_data:
                # 解析菜单数据
                self.parse_user_menu_data(nick_name, menu_data)
                print(f"   ✅ 已处理 {nick_name} 的菜单权限")
            else:
                print(f"   ⚠️  {nick_name} 暂无菜单权限数据")
        
        print(f"🎉 批量处理完成！共处理 {len(self.user_menu_data)} 条菜单权限数据")
        return True
    
    def parse_json_data(self, json_data: List[Dict[str, Any]], parent_path: str = "", level: int = 1) -> None:
        """
        递归解析JSON数据，提取菜单信息
        
        Args:
            json_data: JSON数据列表
            parent_path: 父级菜单路径
            level: 当前层级
        """
        for item in json_data:
            # 获取显示名称和类型
            display_name = item.get('resourcesDisplayName', '')
            resources_type = item.get('resourcesType', 1)
            
            # 构建菜单全路径
            if parent_path:
                menu_path = f"{parent_path} - {display_name}"
            else:
                menu_path = display_name
            
            # 确定类型显示
            type_display = "菜单" if resources_type == 1 else "功能"
            
            # 添加到数据列表
            self.menu_data.append({
                '菜单名称': menu_path,
                '菜单层级': level,
                '类型': type_display
            })
            
            # 递归处理子菜单
            child_resources = item.get('childResources', [])
            if child_resources:
                self.parse_json_data(child_resources, menu_path, level + 1)
    
    def process_json_data(self, data: Any) -> None:
        """
        处理JSON数据，支持多种格式
        
        Args:
            data: JSON数据
        """
        if isinstance(data, list):
            # 如果数据是数组，每个元素都是第一层级的菜单
            self.parse_json_data(data)
        elif isinstance(data, dict):
            # 如果数据是字典，尝试找到包含菜单数据的字段
            if 'data' in data:
                if isinstance(data['data'], list):
                    self.parse_json_data(data['data'])
                else:
                    self.parse_json_data([data['data']])
            elif 'resources' in data:
                if isinstance(data['resources'], list):
                    self.parse_json_data(data['resources'])
                else:
                    self.parse_json_data([data['resources']])
            elif 'childResources' in data:
                # 如果数据本身就是一个菜单对象，直接处理
                self.parse_json_data([data])
            else:
                # 假设数据本身就是菜单
                self.parse_json_data([data])
        else:
            raise ValueError(f"不支持的JSON数据类型: {type(data)}")
    
    def convert_to_excel(self, json_file_path: str = None, json_data: str = None, 
                        use_user_data: bool = False, filename_prefix: str = "") -> str:
        """
        将JSON转换为Excel文件
        
        Args:
            json_file_path: JSON文件路径
            json_data: JSON字符串数据
            use_user_data: 是否使用用户菜单数据
            filename_prefix: 文件名前缀
            
        Returns:
            str: 生成的Excel文件路径
        """
        try:
            # 选择数据源
            if use_user_data:
                if not self.user_menu_data:
                    raise ValueError("用户菜单数据为空，请先调用 batch_process_user_permissions 方法")
                data_to_use = self.user_menu_data
                sheet_name = '用户菜单权限'
                data_count = len(self.user_menu_data)
            else:
                # 解析JSON数据
                if json_file_path:
                    with open(json_file_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                elif json_data:
                    data = json.loads(json_data)
                else:
                    raise ValueError("必须提供JSON文件路径或JSON字符串数据")
                
                # 清空之前的数据
                self.menu_data = []
                
                # 解析数据
                self.process_json_data(data)
                data_to_use = self.menu_data
                sheet_name = '菜单结构'
                data_count = len(self.menu_data)
            
            # 创建DataFrame
            df = pd.DataFrame(data_to_use)
            
            # 生成文件名（yyyyMMddHHmm格式）
            timestamp = datetime.now().strftime('%Y%m%d%H%M')
            if filename_prefix:
                excel_filename = f"{filename_prefix}_{timestamp}.xlsx"
            else:
                excel_filename = f"{timestamp}.xlsx"
            
            # 保存为Excel文件并添加格式
            with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                
                # 获取工作表
                worksheet = writer.sheets[sheet_name]
                
                # 设置列宽
                if use_user_data:
                    worksheet.column_dimensions['A'].width = 20  # 用户名称
                    worksheet.column_dimensions['B'].width = 50  # 菜单名称
                    worksheet.column_dimensions['C'].width = 10  # 菜单层级
                    worksheet.column_dimensions['D'].width = 15  # 类型
                    header_cols = ['A1', 'B1', 'C1', 'D1']
                else:
                    worksheet.column_dimensions['A'].width = 50  # 菜单名称
                    worksheet.column_dimensions['B'].width = 10  # 菜单层级
                    worksheet.column_dimensions['C'].width = 15  # 类型
                    header_cols = ['A1', 'B1', 'C1']
                
                # 设置标题行格式
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                
                # 标题行样式
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # 应用标题行格式
                for col in header_cols:
                    cell = worksheet[col]
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # 冻结标题行（第1行）
                worksheet.freeze_panes = 'A2'
                
                # 设置数据行格式
                data_alignment = Alignment(horizontal="left", vertical="center")
                center_alignment = Alignment(horizontal="center", vertical="center")
                
                # 为不同类型设置不同颜色
                menu_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                function_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                
                # 设置边框
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 应用数据行格式
                for row in range(2, len(df) + 2):  # 从第2行开始（跳过标题行）
                    if use_user_data:
                        # 用户菜单权限数据格式
                        worksheet[f'A{row}'].alignment = center_alignment  # 用户名称
                        worksheet[f'B{row}'].alignment = data_alignment    # 菜单名称
                        worksheet[f'C{row}'].alignment = center_alignment  # 菜单层级
                        worksheet[f'D{row}'].alignment = center_alignment  # 类型
                        
                        # 设置边框
                        for col in ['A', 'B', 'C', 'D']:
                            worksheet[f'{col}{row}'].border = thin_border
                        
                        # 根据类型设置背景色
                        type_cell = worksheet[f'D{row}']
                        if type_cell.value == "菜单":
                            type_cell.fill = menu_fill
                        elif type_cell.value == "功能":
                            type_cell.fill = function_fill
                    else:
                        # 普通菜单数据格式
                        worksheet[f'A{row}'].alignment = data_alignment
                        worksheet[f'B{row}'].alignment = center_alignment
                        worksheet[f'C{row}'].alignment = center_alignment
                        
                        # 设置边框
                        for col in ['A', 'B', 'C']:
                            worksheet[f'{col}{row}'].border = thin_border
                        
                        # 根据类型设置背景色
                        type_cell = worksheet[f'C{row}']
                        if type_cell.value == "菜单":
                            type_cell.fill = menu_fill
                        elif type_cell.value == "功能":
                            type_cell.fill = function_fill
                
                # 如果是用户数据，进行单元格合并
                if use_user_data and len(df) > 0:
                    self._merge_user_cells(worksheet, df)
            
            print(f"✅ Excel文件已生成: {excel_filename}")
            print(f"📊 共处理 {data_count} 条数据")
            
            return excel_filename
            
        except Exception as e:
            print(f"❌ 转换失败: {str(e)}")
            return None
    
    def _merge_user_cells(self, worksheet, df):
        """
        合并相同用户的单元格
        
        Args:
            worksheet: Excel工作表对象
            df: DataFrame数据
        """
        from openpyxl.styles import Alignment
        
        # 按用户名称分组，计算每个用户的行数
        user_groups = df.groupby('用户名称').size().reset_index(name='count')
        
        current_row = 2  # 从第2行开始（跳过标题行）
        
        for _, group in user_groups.iterrows():
            user_name = group['用户名称']
            row_count = group['count']
            
            if row_count > 1:
                # 合并用户名称列的单元格
                start_row = current_row
                end_row = current_row + row_count - 1
                
                # 合并A列（用户名称列）的单元格
                worksheet.merge_cells(f'A{start_row}:A{end_row}')
                
                # 设置合并后单元格的对齐方式
                merged_cell = worksheet[f'A{start_row}']
                merged_cell.alignment = Alignment(
                    horizontal="center", 
                    vertical="center"
                )
            
            current_row += row_count
    
    def print_menu_structure(self, max_items: int = 50) -> None:
        """打印菜单结构（用于调试）"""
        print(f"\n📋 菜单结构预览 (显示前{max_items}项):")
        print("-" * 80)
        
        # 只显示前max_items项，避免输出过多
        display_items = self.menu_data[:max_items]
        for item in display_items:
            indent = "  " * (item['菜单层级'] - 1)
            type_icon = "📁" if item['类型'] == "菜单" else "⚙️"
            print(f"{indent}├─ {type_icon} {item['菜单名称']} (层级: {item['菜单层级']}, 类型: {item['类型']})")
        
        if len(self.menu_data) > max_items:
            print(f"    ... 还有 {len(self.menu_data) - max_items} 项菜单")
        
        # 显示统计信息
        print(f"\n📊 菜单统计:")
        print(f"   总菜单数: {len(self.menu_data)}")
        
        # 按层级统计
        level_counts = {}
        for item in self.menu_data:
            level = item['菜单层级']
            level_counts[level] = level_counts.get(level, 0) + 1
        
        for level in sorted(level_counts.keys()):
            print(f"   第{level}级: {level_counts[level]} 项")
        
        # 按类型统计
        type_counts = {}
        for item in self.menu_data:
            menu_type = item['类型']
            type_counts[menu_type] = type_counts.get(menu_type, 0) + 1
        
        for menu_type, count in type_counts.items():
            print(f"   {menu_type}: {count} 项")


def main():
    """主函数 - 用于JSON文件转Excel"""
    converter = JsonToExcelConverter()
    
    print("🚀 JSON转Excel工具")
    print("=" * 50)
    
    # 示例JSON数据（用于测试）
    sample_json = '''
    [
        {
            "resourcesDisplayName": "系统管理",
            "childResources": [
                {
                    "resourcesDisplayName": "用户管理",
                    "childResources": [
                        {
                            "resourcesDisplayName": "用户列表",
                            "childResources": []
                        },
                        {
                            "resourcesDisplayName": "角色管理",
                            "childResources": []
                        }
                    ]
                },
                {
                    "resourcesDisplayName": "权限管理",
                    "childResources": [
                        {
                            "resourcesDisplayName": "菜单权限",
                            "childResources": []
                        }
                    ]
                }
            ]
        },
        {
            "resourcesDisplayName": "业务管理",
            "childResources": [
                {
                    "resourcesDisplayName": "订单管理",
                    "childResources": []
                }
            ]
        }
    ]
    '''
    
    # 检查是否有命令行参数
    if len(sys.argv) > 1:
        json_file_path = sys.argv[1]
        print(f"📁 使用JSON文件: {json_file_path}")
        excel_file = converter.convert_to_excel(json_file_path=json_file_path)
    else:
        print("📝 使用示例数据")
        excel_file = converter.convert_to_excel(json_data=sample_json)
    
    if excel_file:
        # 显示菜单结构预览
        converter.print_menu_structure()
        
        print(f"\n📄 Excel文件已保存到项目根目录: {excel_file}")
        print("🎉 转换完成！")


def batch_export_user_permissions():
    """批量导出用户菜单权限的主函数"""
    converter = JsonToExcelConverter()
    
    print("🚀 用户菜单权限批量导出工具")
    print("=" * 50)
    
    # 注意：以下参数需要根据实际情况修改
    print("⚠️  请注意：使用前需要在代码中更新以下参数：")
    print("   - token: 认证令牌")
    print("   - userId: 当前用户ID")
    print("   - orgId: 组织ID")
    print("   - eid: 企业ID")
    print("   - uid: 用户唯一标识")
    print("   - 其他必要的请求头和参数")
    print()
    
    # 自定义请求头（请根据实际情况修改）
    custom_headers = {
        'token': 'a7ab82cde0b64c45805d92fa65612c5f',  # 请替换为实际的token
        'userId': '501073',  # 请替换为实际的用户ID
        'orgId': '1000879',  # 请替换为实际的组织ID
        'poit-cloud-org': '1d7d84a6f6b14d6d97f9c7a94813bb22',  # 请替换为实际值
    }
    
    # 自定义请求数据（请根据实际情况修改）
    user_request_data = {
        "eid": "1d7d84a6f6b14d6d97f9c7a94813bb22",  # 请替换为实际的企业ID
        "orgId": "1000879",  # 请替换为实际的组织ID
        "uid": "77c6e33e1b7d4371aecc6477322ff759",  # 请替换为实际的用户唯一标识
        "pageSize": 100,  # 可以根据需要调整页面大小
    }
    
    menu_request_data = {
        "eid": "1d7d84a6f6b14d6d97f9c7a94813bb22",  # 请替换为实际的企业ID
        "orgId": "1000879",  # 请替换为实际的组织ID
        "uid": "77c6e33e1b7d4371aecc6477322ff759",  # 请替换为实际的用户唯一标识
    }
    
    try:
        # 批量处理用户菜单权限
        success = converter.batch_process_user_permissions(
            headers=custom_headers,
            user_request_data=user_request_data,
            menu_request_data=menu_request_data
        )
        
        if success:
            # 导出到Excel
            excel_file = converter.convert_to_excel(
                use_user_data=True, 
                filename_prefix="用户菜单权限"
            )
            
            if excel_file:
                print(f"\n📄 Excel文件已保存到项目根目录: {excel_file}")
                print("🎉 批量导出完成！")
                
                # 显示统计信息
                if converter.user_menu_data:
                    user_count = len(set(item['用户名称'] for item in converter.user_menu_data))
                    print(f"\n📊 导出统计:")
                    print(f"   用户总数: {user_count}")
                    print(f"   菜单权限总数: {len(converter.user_menu_data)}")
            else:
                print("❌ Excel文件生成失败")
        else:
            print("❌ 批量处理失败")
            
    except Exception as e:
        print(f"❌ 处理过程中发生错误: {str(e)}")


if __name__ == "__main__":
    # 检查命令行参数来决定运行哪个功能
    if len(sys.argv) > 1 and sys.argv[1] == "--batch-users":
        batch_export_user_permissions()
    else:
        main() 