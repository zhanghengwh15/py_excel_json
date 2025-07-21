#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
用户角色权限导出工具
功能：获取用户列表，关联角色权限，导出用户菜单权限详情到Excel
"""

import json
import pandas as pd
import os
import time
import requests
from datetime import datetime
from typing import List, Dict, Any, Set
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class UserRolePermissionExporter:
    """用户角色权限导出器"""
    
    def __init__(self):
        """初始化导出器"""
        self.user_menu_data = []
        self.id_map = {}
        self.role_permissions = {}
        
    def load_id_map(self, id_map_file: str = "id_map.json") -> bool:
        """
        加载ID映射文件
        
        Args:
            id_map_file: ID映射文件路径
            
        Returns:
            bool: 是否加载成功
        """
        try:
            with open(id_map_file, 'r', encoding='utf-8') as f:
                id_map_list = json.load(f)
            
            # 转换为字典格式，key为resourcesId，value为完整信息
            for item in id_map_list:
                self.id_map[item['resourcesId']] = item
            
            print(f"✅ 成功加载ID映射文件，包含 {len(self.id_map)} 个资源")
            return True
            
        except FileNotFoundError:
            print(f"❌ 找不到ID映射文件: {id_map_file}")
            return False
        except json.JSONDecodeError as e:
            print(f"❌ ID映射文件格式错误: {e}")
            return False
    
    def load_role_permissions(self, role_dir: str = "role") -> bool:
        """
        加载角色权限文件
        
        Args:
            role_dir: 角色权限文件目录
            
        Returns:
            bool: 是否加载成功
        """
        if not os.path.exists(role_dir):
            print(f"❌ 角色权限目录不存在: {role_dir}")
            return False
        
        loaded_count = 0
        for filename in os.listdir(role_dir):
            if filename.endswith('.json'):
                role_id = filename.replace('.json', '')
                file_path = os.path.join(role_dir, filename)
                
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        permission_data = json.load(f)
                    
                    self.role_permissions[role_id] = permission_data
                    loaded_count += 1
                    
                except Exception as e:
                    print(f"⚠️ 加载角色权限文件 {filename} 失败: {e}")
        
        print(f"✅ 成功加载 {loaded_count} 个角色权限文件")
        return loaded_count > 0
    
    def get_user_list(self, base_url: str = "https://cloudsy.shede.com.cn", 
                     headers: Dict[str, str] = None, 
                     request_data: Dict[str, Any] = None) -> List[Dict[str, Any]]:
        """
        获取用户列表（支持分页）
        
        Args:
            base_url: 基础URL
            headers: 请求头
            request_data: 请求数据
            
        Returns:
            List[Dict]: 用户列表，包含 nickName、userId、roleIdList
        """
        url = f"{base_url}/api/poit-cloud-platform/user/ent/user/page"
        
        # 默认请求头
        default_headers = {
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'zh-cn',
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'Content-Type': 'application/json; charset=UTF-8',
            'Origin': base_url,
            'Pragma': 'no-cache',
            'Referer': f'{base_url}/cloud/cloud_config/userManagement',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-origin',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36',
            'X-Requested-With': 'XMLHttpRequest',
            'loglevel': 'debug',
            'menusUri': '/userManagement',
            'orgId': '1000879',
            'poit-cloud-org': '1d7d84a6f6b14d6d97f9c7a94813bb22',
            'poit-cloud-src-client': 'cloud',
            'sec-ch-ua': '"Not)A;Brand";v="8", "Chromium";v="138", "Google Chrome";v="138"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"macOS"',
            'token': 'a7ab82cde0b64c45805d92fa65612c5f',
            'userId': '501073',
            'x-poit-tif-submit': '1675135120_1753070008803'
        }
        
        # 合并自定义请求头
        if headers:
            default_headers.update(headers)
        
        # 默认请求数据
        default_data = {
            "eid": "1d7d84a6f6b14d6d97f9c7a94813bb22",
            "operateUserId": "501073",
            "orgId": "1000879",
            "uid": "77c6e33e1b7d4371aecc6477322ff759",
            "appVersion": "1.0",
            "keyword": "",
            "combineStatus": None,
            "pageSize": 100,
            "pageNum": 1,
            "positionIdList": [],
            "roleIdList": [],
            "factoryIdList": [],
            "departmentIdList": []
        }
        
        # 合并自定义请求数据
        if request_data:
            default_data.update(request_data)
        
        all_users = []
        page_num = 1
        
        try:
            while True:
                # 更新页码
                default_data['pageNum'] = page_num
                
                print(f"📄 正在获取第 {page_num} 页用户数据...")
                
                response = requests.post(url, headers=default_headers, json=default_data)
                response.raise_for_status()
                
                result = response.json()
                
                # 检查响应数据
                if 'data' not in result:
                    print("❌ 响应数据格式不正确，缺少data字段")
                    break
                
                # 如果data为空，说明已经获取完所有数据
                if not result['data'] or len(result['data']) == 0:
                    print(f"✅ 已获取完所有用户数据，共 {len(all_users)} 个用户")
                    break
                
                # 提取用户数据
                if isinstance(result['data'], list):
                    page_users = []
                    for user in result['data']:
                        page_users.append({
                            'nickName': user.get('nickName', ''),
                            'userId': user.get('userId', ''),
                            'openUserId': user.get('openUserId', ''),
                            'roleIdList': user.get('roleIdList', [])  # 角色ID列表
                        })
                    
                    all_users.extend(page_users)
                    print(f"   ✅ 第 {page_num} 页获取到 {len(page_users)} 个用户")
                    
                    # 如果当前页的用户数量小于pageSize，说明已经是最后一页
                    if len(page_users) < default_data['pageSize']:
                        print(f"✅ 已获取完所有用户数据，共 {len(all_users)} 个用户")
                        break
                else:
                    print("❌ 响应data字段不是数组格式")
                    break
                
                # 页码递增
                page_num += 1
                
                # 请求间隔200ms
                time.sleep(0.2)
                
        except requests.exceptions.RequestException as e:
            print(f"❌ 获取用户列表失败: {str(e)}")
            return all_users  # 返回已获取的数据
        
        return all_users
    
    def process_user_permissions(self, users: List[Dict[str, Any]]) -> None:
        """
        处理用户权限数据
        
        Args:
            users: 用户列表
        """
        print("🔄 开始处理用户权限数据...")
        
        for i, user in enumerate(users, 1):
            nick_name = user['nickName']
            role_id_list = user.get('roleIdList', [])
            
            print(f"📊 正在处理用户 {i}/{len(users)}: {nick_name}")
            
            # 收集该用户的所有权限ID
            user_permission_ids = set()
            
            # 遍历用户的角色ID列表
            for role_id in role_id_list:
                role_id_str = str(role_id)
                if role_id_str in self.role_permissions:
                    # 获取该角色的权限ID列表
                    role_permission_ids = self.role_permissions[role_id_str]
                    if isinstance(role_permission_ids, list):
                        user_permission_ids.update(role_permission_ids)
                    else:
                        print(f"   ⚠️ 角色 {role_id} 的权限数据格式不正确")
                else:
                    print(f"   ⚠️ 找不到角色 {role_id} 的权限文件")
            
            # 根据权限ID获取菜单信息
            for permission_id in user_permission_ids:
                if permission_id in self.id_map:
                    menu_info = self.id_map[permission_id]
                    
                    # 添加到用户菜单数据列表
                    self.user_menu_data.append({
                        '用户名称': nick_name,
                        '菜单名称': menu_info['path'],
                        '菜单层级': menu_info['level'],
                        '资源类型': menu_info['resourcesType']
                    })
                else:
                    print(f"   ⚠️ 找不到资源ID {permission_id} 的映射信息")
        
        print(f"✅ 用户权限数据处理完成，共生成 {len(self.user_menu_data)} 条记录")
    
    def export_to_excel(self, filename_prefix: str = "用户角色权限") -> str:
        """
        导出到Excel文件
        
        Args:
            filename_prefix: 文件名前缀
            
        Returns:
            str: 生成的Excel文件路径
        """
        if not self.user_menu_data:
            print("❌ 没有数据可导出")
            return None
        
        try:
            # 创建DataFrame
            df = pd.DataFrame(self.user_menu_data)
            
            # 按用户名称、菜单层级、排序值排序
            df = df.sort_values(['用户名称', '菜单层级', '排序值', '菜单名称'])
            
            # 生成文件名
            timestamp = datetime.now().strftime('%Y%m%d%H%M')
            excel_filename = f"{filename_prefix}_{timestamp}.xlsx"
            
            # 保存为Excel文件并添加格式
            with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='用户角色权限')
                
                # 获取工作表
                worksheet = writer.sheets['用户角色权限']
                
                # 设置列宽
                worksheet.column_dimensions['A'].width = 20  # 用户名称
                worksheet.column_dimensions['B'].width = 50  # 菜单名称
                worksheet.column_dimensions['C'].width = 10  # 菜单层级
                worksheet.column_dimensions['D'].width = 15  # 资源类型
                
                # 设置标题行格式
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # 应用标题行格式
                for col in ['A1', 'B1', 'C1', 'D1']:
                    cell = worksheet[col]
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # 冻结标题行
                worksheet.freeze_panes = 'A2'
                
                # 设置数据行格式
                data_alignment = Alignment(horizontal="left", vertical="center")
                center_alignment = Alignment(horizontal="center", vertical="center")
                
                # 设置边框
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 应用数据行格式
                for row in range(2, len(df) + 2):
                    # 设置对齐方式
                    worksheet[f'A{row}'].alignment = center_alignment  # 用户名称
                    worksheet[f'B{row}'].alignment = data_alignment    # 菜单名称
                    worksheet[f'C{row}'].alignment = center_alignment  # 菜单层级
                    worksheet[f'D{row}'].alignment = center_alignment  # 资源类型
                    
                    # 设置边框
                    for col in ['A', 'B', 'C', 'D']:
                        worksheet[f'{col}{row}'].border = thin_border
                
                # 合并相同用户的单元格
                self._merge_user_cells(worksheet, df)
            
            print(f"✅ Excel文件已生成: {excel_filename}")
            print(f"📊 共处理 {len(self.user_menu_data)} 条数据")
            
            return excel_filename
            
        except Exception as e:
            print(f"❌ 导出失败: {str(e)}")
            return None
    
    def _merge_user_cells(self, worksheet, df):
        """
        合并相同用户的单元格
        
        Args:
            worksheet: Excel工作表对象
            df: DataFrame数据
        """
        # 按用户名称分组，计算每个用户的行数
        user_groups = df.groupby('用户名称').size().reset_index(name='count')
        
        current_row = 2  # 从第2行开始（跳过标题行）
        
        for _, group in user_groups.iterrows():
            user_name = group['用户名称']
            row_count = group['count']
            
            if row_count > 1:
                # 合并用户名称列的单元格
                start_row = current_row
                end_row = current_row + row_count - 1
                
                # 合并A列（用户名称列）的单元格
                worksheet.merge_cells(f'A{start_row}:A{end_row}')
                
                # 设置合并后单元格的对齐方式
                merged_cell = worksheet[f'A{start_row}']
                merged_cell.alignment = Alignment(
                    horizontal="center", 
                    vertical="center"
                )
            
            current_row += row_count
    
    def run_export(self, headers: Dict[str, str] = None, 
                  request_data: Dict[str, Any] = None) -> bool:
        """
        执行完整的导出流程
        
        Args:
            headers: 自定义请求头
            request_data: 自定义请求数据
            
        Returns:
            bool: 是否成功
        """
        print("🚀 开始用户角色权限导出任务")
        print("=" * 60)
        
        # 1. 加载ID映射文件
        print("📋 步骤1: 加载ID映射文件...")
        if not self.load_id_map():
            return False
        
        # 2. 加载角色权限文件
        print("📋 步骤2: 加载角色权限文件...")
        if not self.load_role_permissions():
            return False
        
        # 3. 获取用户列表
        print("📋 步骤3: 获取用户列表...")
        users = self.get_user_list(headers=headers, request_data=request_data)
        
        if not users:
            print("❌ 未获取到用户数据")
            return False
        
        print(f"✅ 获取到 {len(users)} 个用户")
        
        # 4. 处理用户权限数据
        print("📋 步骤4: 处理用户权限数据...")
        self.process_user_permissions(users)
        
        # 5. 导出到Excel
        print("📋 步骤5: 导出到Excel...")
        excel_file = self.export_to_excel()
        
        if excel_file:
            print("=" * 60)
            print(f"🎉 导出完成！文件保存为: {excel_file}")
            
            # 显示统计信息
            user_count = len(set(item['用户名称'] for item in self.user_menu_data))
            print(f"\n📊 导出统计:")
            print(f"   用户总数: {user_count}")
            print(f"   权限记录总数: {len(self.user_menu_data)}")
            
            return True
        else:
            print("❌ 导出失败")
            return False


def main():
    """主函数"""
    exporter = UserRolePermissionExporter()
    
    # 自定义请求头（请根据实际情况修改）
    custom_headers = {
        'token': 'a7ab82cde0b64c45805d92fa65612c5f',  # 请替换为实际的token
        'userId': '501073',  # 请替换为实际的用户ID
        'orgId': '1000879',  # 请替换为实际的组织ID
        'poit-cloud-org': '1d7d84a6f6b14d6d97f9c7a94813bb22',  # 请替换为实际值
    }
    
    # 自定义请求数据（请根据实际情况修改）
    custom_request_data = {
        "eid": "1d7d84a6f6b14d6d97f9c7a94813bb22",  # 请替换为实际的企业ID
        "orgId": "1000879",  # 请替换为实际的组织ID
        "uid": "77c6e33e1b7d4371aecc6477322ff759",  # 请替换为实际的用户唯一标识
        "pageSize": 100,  # 可以根据需要调整页面大小
    }
    
    try:
        success = exporter.run_export(
            headers=custom_headers,
            request_data=custom_request_data
        )
        
        if not success:
            print("❌ 导出任务失败")
            
    except Exception as e:
        print(f"❌ 执行过程中发生错误: {str(e)}")


if __name__ == "__main__":
    main() 