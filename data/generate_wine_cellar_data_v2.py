#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
酒鬼窖池历史数据生成器 V2
功能：根据Excel模板格式生成酒鬼窖池历史数据
作者：AI Assistant
日期：2024
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import random
import math
import os
from typing import List, Dict, Any


class WineCellarDataGeneratorV2:
    """酒鬼窖池历史数据生成器 V2"""
    
    def __init__(self):
        self.cellar_ids = [f"41.{i:03d}" for i in range(1, 12)]  # 窖池ID: 41.001-41.011
        self.data_count = 100  # 每个sheet的数据量改为
        
    def generate_cellar_code(self, cellar_id: str) -> str:
        """生成窖池编码（格式：41.001、41.002等）"""
        return cellar_id
    
    def generate_batch_number(self, year: int, batch: int) -> str:
        """生成轮次编号（格式：2025-1）"""
        return f"{year}-{batch}"
    
    def generate_random_date(self, start_year: int = 2000, end_year: int = 2025) -> str:
        """生成随机日期（格式：2023/11/11），时间范围1970-2025年"""
        start_date = datetime(start_year, 1, 1)
        end_date = datetime(end_year, 12, 31)
        time_between_dates = end_date - start_date
        days_between_dates = time_between_dates.days
        random_number_of_days = random.randrange(days_between_dates)
        random_date = start_date + timedelta(days=random_number_of_days)
        return random_date.strftime('%Y/%m/%d')
    
    def generate_random_date_by_batch(self, year: int, batch: int) -> str:
        """根据轮次生成对应月份内的随机日期，batch 1-12 表示月份"""
        # 确保月份合法
        month = batch if 1 <= batch <= 12 else 1
        # 月初
        start_date = datetime(year, month, 1)
        # 月末
        if month == 12:
            end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1) - timedelta(days=1)
        # 随机天数
        days_between = (end_date - start_date).days
        random_days = random.randint(0, days_between)
        random_date = start_date + timedelta(days=random_days)
        return random_date.strftime('%Y/%m/%d')
    
    def generate_sheet1_data(self) -> List[Dict[str, Any]]:
        """生成第一个sheet的数据（投料耗用数据）"""
        data = []
        
        # 生成与数据量对应的唯一(year, batch)组合列表
        all_pairs = [(year, batch) for year in range(2000, 2026) for batch in range(1, 13)]
        random.shuffle(all_pairs)
        if self.data_count <= len(all_pairs):
            selected_pairs = all_pairs[:self.data_count]
        else:
            repeats = math.ceil(self.data_count / len(all_pairs))
            selected_pairs = (all_pairs * repeats)[:self.data_count]
        
        # 为每个(year, batch)组合分配窖池号，确保sheet1和sheet2的一致性
        cellar_batch_combinations = []
        for i, (year, batch) in enumerate(selected_pairs):
            cellar_id = self.cellar_ids[i % len(self.cellar_ids)]
            cellar_batch_combinations.append((cellar_id, year, batch))
        
        # 保存供sheet2使用
        self.cellar_batch_combinations = cellar_batch_combinations
        
        for i, (cellar_id, year, batch) in enumerate(cellar_batch_combinations):
            record = {
                '*窖池编号': self.generate_cellar_code(cellar_id),
                '*投入日期': self.generate_random_date_by_batch(year, batch),
                '*轮次': self.generate_batch_number(year, batch),
                '*发酵期(天)': random.randint(30, 90),
                '糖化粮(投入)': random.randint(100, 500),
                '大曲(投入)': random.randint(50, 200),
                '熟谷壳(投入)': random.randint(20, 100),
                '状态': random.choice(['投入', '耗用', '发酵中']),
                '糖化粮(耗用)': random.randint(80, 450),
                '大曲(耗用)': random.randint(40, 180),
                '熟谷壳(耗用)': random.randint(15, 90)
            }
            data.append(record)
        
        return data
    
    def generate_sheet2_data(self) -> List[Dict[str, Any]]:
        """生成第二个sheet的数据（等级酒生产记录）"""
        data = []
        
        # 使用与sheet1相同的窖池号+轮次组合
        if not hasattr(self, 'cellar_batch_combinations'):
            raise ValueError("Cellar-batch combinations for sheet1 not generated yet")
        combinations_list = self.cellar_batch_combinations.copy()
        # 确保使用相同的组合，不重复，不遗漏
        if self.data_count <= len(combinations_list):
            selected_combinations = combinations_list[:self.data_count]
        else:
            repeats = math.ceil(self.data_count / len(combinations_list))
            selected_combinations = (combinations_list * repeats)[:self.data_count]
        
        for i, (cellar_id, year, batch) in enumerate(selected_combinations):
            # 生成原酒分类数据
            raw_wine_z1 = random.randint(10, 50)
            raw_wine_z2 = random.randint(10, 50)
            raw_wine_d1 = random.randint(5, 30)
            raw_wine_d2 = random.randint(10, 50)
            raw_wine_d3 = random.randint(10, 50)
            raw_wine_g = random.randint(10, 50)
            
            # 计算总产量和出酒率
            total_yield = raw_wine_z1 + raw_wine_z2 + raw_wine_d1 + raw_wine_d2 + raw_wine_d3 + raw_wine_g
            yield_rate = round(random.uniform(0.6, 0.9), 2)
            
            record = {
                '*窖池编号': self.generate_cellar_code(cellar_id),
                '*日期': self.generate_random_date_by_batch(year, batch),
                '*轮次': self.generate_batch_number(year, batch),
                '原酒Z1': raw_wine_z1,
                '原酒Z2': raw_wine_z2,
                '原酒D1': raw_wine_d1,
                '原酒D2': raw_wine_d2,
                '原酒D3': raw_wine_d3,
                '原酒G': raw_wine_g,
                '未折度产量 (kg)': total_yield,
                '未折度出酒率 (%)': yield_rate * 100
            }
            data.append(record)
        
        return data
    
    def generate_sheet3_data(self) -> List[Dict[str, Any]]:
        """生成第三个sheet的数据（入窖出窖糟化验记录）"""
        data = []
        
        for _ in range(self.data_count):
            cellar_id = random.choice(self.cellar_ids)
            
            # 出池化验数据
            out_acidity = round(random.uniform(3.0, 8.0), 2)
            out_starch = round(random.uniform(0.5, 3.0), 2)
            out_moisture = round(random.uniform(40.0, 60.0), 2)
            
            # 入池化验数据
            in_acidity = round(random.uniform(2.0, 6.0), 2)
            in_starch = round(random.uniform(15.0, 25.0), 2)
            in_moisture = round(random.uniform(45.0, 65.0), 2)
            
            record = {
                '*窖池编号': self.generate_cellar_code(cellar_id),
                '*日期': self.generate_random_date(),
                '出池化验数据_酸度': out_acidity,
                '出池化验数据_淀粉': out_starch,
                '出池化验数据_水分': out_moisture,
                '入池化验数据_酸度': in_acidity,
                '入池化验数据_淀粉': in_starch,
                '入池化验数据_水分': in_moisture
            }
            data.append(record)
        
        return data
    
    def generate_all_data(self) -> Dict[str, pd.DataFrame]:
        """生成所有数据"""
        print("🔄 正在生成投料耗用数据...")
        sheet1_data = self.generate_sheet1_data()
        df1 = pd.DataFrame(sheet1_data)
        
        print("🔄 正在生成等级酒生产记录...")
        sheet2_data = self.generate_sheet2_data()
        df2 = pd.DataFrame(sheet2_data)
        
        print("🔄 正在生成入窖出窖糟化验记录...")
        sheet3_data = self.generate_sheet3_data()
        df3 = pd.DataFrame(sheet3_data)
        
        # 保存到all_data
        all_data = {
            "投料耗用": df1,
            "等级酒生产记录": df2,
            "入窖出窖糟化验记录": df3
        }
        
        print(f"✅ 数据生成完成，共 {len(df1) + len(df2) + len(df3)} 条记录")
        
        return all_data
    
    def save_to_excel(self, data: Dict[str, pd.DataFrame], filename: str = None) -> str:
        """保存数据到Excel文件"""
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d%H%M')
            filename = f"酒鬼窖池历史数据导入_{timestamp}.xlsx"
        
        filepath = os.path.join('data', filename)
        
        print(f"💾 正在保存数据到: {filepath}")
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            for sheet_name, df in data.items():
                # 保存到Excel
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 获取工作表进行格式设置
                worksheet = writer.sheets[sheet_name]
                
                # 设置列宽
                for col_num, column in enumerate(df.columns, 1):
                    col_letter = chr(64 + col_num)  # A, B, C...
                    if col_num > 26:  # 处理超过26列的情况
                        col_letter = 'A' + chr(64 + col_num - 26)
                    
                    # 根据列内容设置合适的宽度
                    if '日期' in column or '轮次' in column:
                        worksheet.column_dimensions[col_letter].width = 15
                    elif '编号' in column:
                        worksheet.column_dimensions[col_letter].width = 12
                    elif '率' in column or '期' in column:
                        worksheet.column_dimensions[col_letter].width = 12
                    else:
                        worksheet.column_dimensions[col_letter].width = 10
                
                # 设置标题行格式
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                
                # 标题行样式
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # 根据不同的sheet设置不同的标题格式
                if sheet_name == "等级酒生产记录":
                    # 插入合并单元格的标题行
                    worksheet.insert_rows(1)
                    
                    # 设置第一行标题（合并单元格）
                    worksheet['A1'] = '窖池编号'
                    worksheet['B1'] = '日期'
                    worksheet['C1'] = '轮次'
                    worksheet['D1'] = '原酒'  # 合并D1:I1
                    worksheet['J1'] = '未折度产量 (kg)'
                    worksheet['K1'] = '未折度出酒率 (%)'
                    
                    # 设置第二行子标题
                    worksheet['A2'] = '*窖池编号'
                    worksheet['B2'] = '*日期'
                    worksheet['C2'] = '*轮次'
                    worksheet['D2'] = '原酒Z1'
                    worksheet['E2'] = '原酒Z2'
                    worksheet['F2'] = '原酒D1'
                    worksheet['G2'] = '原酒D2'
                    worksheet['H2'] = '原酒D3'
                    worksheet['I2'] = '原酒G'
                    worksheet['J2'] = '未折度产量 (kg)'
                    worksheet['K2'] = '未折度出酒率 (%)'
                    
                    # 合并单元格
                    worksheet.merge_cells('D1:I1')
                    
                    # 应用标题行格式
                    for col in ['A1', 'B1', 'C1', 'D1', 'J1', 'K1']:
                        cell = worksheet[col]
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    for col in ['A2', 'B2', 'C2', 'D2', 'E2', 'F2', 'G2', 'H2', 'I2', 'J2', 'K2']:
                        cell = worksheet[col]
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    # 冻结前两行
                    worksheet.freeze_panes = 'A3'
                    
                elif sheet_name == "入窖出窖糟化验记录":
                    # 插入合并单元格的标题行
                    worksheet.insert_rows(1)
                    
                    # 设置第一行标题（合并单元格）
                    worksheet['A1'] = '窖池编号'
                    worksheet['B1'] = '日期'
                    worksheet['C1'] = '出池化验数据'  # 合并C1:E1
                    worksheet['F1'] = '入池化验数据'  # 合并F1:H1
                    
                    # 设置第二行子标题
                    worksheet['A2'] = '*窖池编号'
                    worksheet['B2'] = '*日期'
                    worksheet['C2'] = '酸度'
                    worksheet['D2'] = '淀粉'
                    worksheet['E2'] = '水分'
                    worksheet['F2'] = '酸度'
                    worksheet['G2'] = '淀粉'
                    worksheet['H2'] = '水分'
                    
                    # 合并单元格
                    worksheet.merge_cells('C1:E1')
                    worksheet.merge_cells('F1:H1')
                    
                    # 应用标题行格式
                    for col in ['A1', 'B1', 'C1', 'F1']:
                        cell = worksheet[col]
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    for col in ['A2', 'B2', 'C2', 'D2', 'E2', 'F2', 'G2', 'H2']:
                        cell = worksheet[col]
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    # 冻结前两行
                    worksheet.freeze_panes = 'A3'
                    
                else:
                    # 投料耗用数据使用普通标题行
                    # 应用标题行格式
                    for col_num in range(1, len(df.columns) + 1):
                        col_letter = chr(64 + col_num)
                        if col_num > 26:
                            col_letter = 'A' + chr(64 + col_num - 26)
                        
                        cell = worksheet[f'{col_letter}1']
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    # 冻结标题行
                    worksheet.freeze_panes = 'A2'
                
                # 设置数据行格式
                data_alignment = Alignment(horizontal="center", vertical="center")
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 应用数据行格式（只对前1000行应用格式，避免文件过大）
                start_row = 3 if sheet_name in ["等级酒生产记录", "入窖出窖糟化验记录"] else 2
                max_format_rows = min(1000, len(df))  # 只对前1000行应用格式
                
                for row in range(start_row, start_row + max_format_rows):
                    for col_num in range(1, len(df.columns) + 1):
                        col_letter = chr(64 + col_num)
                        if col_num > 26:
                            col_letter = 'A' + chr(64 + col_num - 26)
                        
                        cell = worksheet[f'{col_letter}{row}']
                        cell.alignment = data_alignment
                        cell.border = thin_border
        
        print(f"✅ Excel文件保存成功: {filepath}")
        return filepath
    
    def print_data_summary(self, data: Dict[str, pd.DataFrame]) -> None:
        """打印数据摘要"""
        print("\n📊 数据生成摘要:")
        print("=" * 50)
        
        total_records = 0
        
        for sheet_name, df in data.items():
            record_count = len(df)
            total_records += record_count
            print(f"📋 {sheet_name}: {record_count:,} 条记录")
        
        print(f"\n📈 总计: {len(data)} 个工作表，{total_records:,} 条记录")
        
        # 显示数据示例
        if data:
            first_sheet = list(data.keys())[0]
            first_df = data[first_sheet]
            print(f"\n📝 {first_sheet} 数据示例:")
            print(first_df.head(3).to_string(index=False))


def main():
    """主函数"""
    print("🍷 酒鬼窖池历史数据生成器 V2")
    print("=" * 50)
    
    # 创建数据生成器
    generator = WineCellarDataGeneratorV2()
    
    # 生成所有数据
    print("🚀 开始生成数据...")
    all_data = generator.generate_all_data()
    
    # 保存到Excel
    excel_file = generator.save_to_excel(all_data)
    
    # 打印摘要
    generator.print_data_summary(all_data)
    
    print(f"\n🎉 数据生成完成！")
    print(f"📁 文件位置: {excel_file}")


if __name__ == "__main__":
    main() 